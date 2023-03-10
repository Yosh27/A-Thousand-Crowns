import openpyxl

# RGBA map of the Smiley icon. Transparent pixels are marked as none for ease/coder's sanity
SMILEYICON = [[None, None, None, None, None, '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', None, None, None, None, None],
              [None, None, None, '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', None, None, None],
              [None, None, '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', None, None],
              [None, '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', None],
              [None, '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', None],
              ['[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[0, 0, 0, 255]', '[0, 0, 0, 255]', '[0, 0, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[0, 0, 0, 255]', '[0, 0, 0, 255]', '[0, 0, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]'],
              ['[255, 255, 0, 255]', '[255, 255, 0, 255]', '[0, 0, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[0, 0, 0, 255]', '[255, 255, 0, 255]', '[0, 0, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[0, 0, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]'],
              ['[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]'],
              ['[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]'],
              ['[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]'],
              [None, '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[0, 0, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[0, 0, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', None],
              [None, '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[0, 0, 0, 255]', '[0, 0, 0, 255]', '[0, 0, 0, 255]', '[0, 0, 0, 255]', '[0, 0, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', None],
              [None, None, '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[0, 0, 0, 255]', '[0, 0, 0, 255]', '[0, 0, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', None, None],
              [None, None, None, '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', None, None, None],
              [None, None, None, None, None, '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', '[255, 255, 0, 255]', None, None, None, None, None]]

'''
    The data in SMILEYDATA represents the length of rows, length of columns, center row, and center column respectively
    
    For even and odd lengths of rows or columns, the central row/column will be the (length // 2 + 1)
    
    The center row/column for even numbers is really (length / 2), but it's works out better with the equation above (I think)
'''
SMILEYDATA = (15, 15, 8, 8)
HYPOTHETICALEVENDATA = (10, 10, 6, 6)

def iconsOverlap(sheet, x, y, iconData):
    for row in range(y, y + iconData[0]):
        for column in range(x, x + iconData[1]):
            if sheet.cell(row, column).value != None:
                print(f'Icon overlaps another at cell {row} {column}')
                return True
    return False


def placeIcon(sheet, x, y, icon, iconData):
    x -= (iconData[1] - iconData[3])
    y -= (iconData[0] - iconData[2])
    
    if not iconsOverlap(sheet, x, y, iconData):
        i = 0
        j = 0
        for row in range(y, y + iconData[0]):
            for column in range(x, x + iconData[1]):
                if icon[i][j] != None:
                   cell = sheet.cell(row, column)
                   cell.value = icon[i][j]
                j += 1
            j = 0
            i += 1
        workBook.save('C:\\Users...foo.xlsx') # Enter a path to a valid excel file here. Also, the excel sheet can't be opened while saving takes place
        print('Icon recorded successfully!')


if __name__=='__main__':
    workBook = openpyxl.load_workbook('C:\\Users...foo.xlsx') # Enter a path to a valid excel file here
    sheet = workBook.active

    placeIcon(sheet, 600, 400, SMILEYICON, SMILEYDATA)